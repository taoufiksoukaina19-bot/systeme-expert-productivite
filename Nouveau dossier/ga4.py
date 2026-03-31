import openpyxl
import pandas as pd
import os
import traceback

# Définition des étapes
STAGES = ["découpe", "placage de chant", "perçage", "montage", "conditionnement"]

def get_stage(description):
    """Retourne l'étape à partir de la description du composant."""
    if description is None:
        return "montage"
    desc = str(description).upper()
    if "PANNEAU" in desc or "PVC" in desc:
        return "découpe"
    if "BANDE DE CHANT" in desc or "COLLE DE CHANT" in desc:
        return "placage de chant"
    # Mots clés pour le conditionnement (emballage)
    packaging = ["SACHET", "POIGNEE", "CARTON D'EMBALLAGE", "DOUBLE HOOK",
                 "VIS POIGNEE", "PALETTE", "AGRAFES"]
    if any(kw in desc for kw in packaging):
        return "conditionnement"
    # Par défaut, tout le reste va en montage
    return "montage"

def main():
    try:
        # Nom du fichier source
        fichier_source = "Nomenclature CEJID 2026.xlsx"
        
        print(f"📁 Répertoire de travail : {os.getcwd()}")
        
        # Vérifier si le fichier existe
        if not os.path.exists(fichier_source):
            print(f"❌ Erreur : Le fichier '{fichier_source}' n'existe pas.")
            # Lister les fichiers Excel dans le dossier
            fichiers_excel = [f for f in os.listdir() if f.endswith('.xlsx')]
            print(f"📄 Fichiers Excel trouvés : {fichiers_excel}")
            return
        
        print(f"📂 Chargement du fichier : {fichier_source}")
        
        # Charger le classeur source avec évaluation des formules
        wb = openpyxl.load_workbook(fichier_source, data_only=True)
        
        print(f"📑 Feuilles disponibles : {wb.sheetnames}")
        
        # Nom de la feuille à traiter
        TARGET_SHEET = "PVC 2026"
        
        if TARGET_SHEET not in wb.sheetnames:
            print(f"❌ Erreur : La feuille '{TARGET_SHEET}' n'existe pas.")
            return
        
        sheet = wb[TARGET_SHEET]
        print(f"✅ Traitement de la feuille : {TARGET_SHEET}")
        
        all_rows = []
        current_product = None
        row_count = 0

        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if len(row) < 6:
                continue
            
            # Colonnes : Reference, Designation, Composant, Désignation Componsant, Quantité, Unité
            ref, designation, composant, desc_composant, quantity, unit = row[:6]

            # Mise à jour du produit courant (propagation)
            if designation and str(designation).strip():
                current_product = str(designation).strip()
                print(f"   Produit trouvé ligne {row_num}: {current_product[:50]}...")

            # Ignorer les lignes sans produit
            if not current_product:
                continue

            # Ignorer les composants MOD / ZMOD
            if composant and str(composant).upper() in ["MOD", "ZMOD"]:
                continue

            # Ignorer les lignes sans description de composant
            if not desc_composant:
                continue

            # Déterminer l'étape (gamme)
            stage = get_stage(desc_composant)
            
            # Gérer les formules Excel dans la quantité
            if isinstance(quantity, (int, float)):
                qty_value = quantity
            elif quantity is None:
                qty_value = 0
            else:
                try:
                    qty_value = float(quantity) if quantity else 0
                except (ValueError, TypeError):
                    qty_value = 0

            all_rows.append({
                "Produit": current_product,
                "Gamme": stage,
                "Code Composant": str(composant) if composant else "",
                "Désignation Composant": str(desc_composant),
                "Quantité": qty_value,
                "Unité": unit if unit else ""
            })
            row_count += 1

        print(f"\n📊 Données extraites : {row_count} lignes")

        if not all_rows:
            print("⚠️ Aucune donnée trouvée dans la feuille.")
            return

        # Créer le DataFrame
        df = pd.DataFrame(all_rows)
        
        print("📝 Création du fichier Excel...")
        
        # Créer un fichier Excel avec plusieurs onglets
        output_file = "Nomenclature_PVC_2026_par_etape.xlsx"
        
        # Utiliser ExcelWriter avec un contexte
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Onglet 1 : Toutes les données
                df.to_excel(writer, sheet_name='Toutes_les_donnees', index=False)
                print("   ✅ Onglet 'Toutes_les_donnees' créé")
                
                # Onglet 2 : Données par étape
                for stage in df['Gamme'].unique():
                    df_stage = df[df['Gamme'] == stage]
                    sheet_name = f'Etape_{stage}'[:31]  # Excel limite à 31 caractères
                    df_stage.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"   ✅ Onglet '{sheet_name}' créé ({len(df_stage)} lignes)")
                
                # Onglet 3 : Résumé par produit
                resume_produit = df.groupby('Produit').agg({
                    'Code Composant': 'count',
                    'Quantité': 'sum'
                }).rename(columns={'Code Composant': 'Nombre_de_composants', 'Quantité': 'Quantite_totale'})
                resume_produit.to_excel(writer, sheet_name='Resume_par_produit')
                print(f"   ✅ Onglet 'Resume_par_produit' créé ({len(resume_produit)} produits)")
                
                # Onglet 4 : Résumé par étape
                resume_etape = df.groupby('Gamme').size().to_frame('Nombre_de_composants')
                resume_etape.to_excel(writer, sheet_name='Resume_par_etape')
                print("   ✅ Onglet 'Resume_par_etape' créé")
            
            print(f"\n✅ SUCCÈS ! Fichier généré : {output_file}")
            print(f"📁 Chemin complet : {os.path.abspath(output_file)}")
            
            # Vérifier que le fichier a bien été créé
            if os.path.exists(output_file):
                taille = os.path.getsize(output_file)
                print(f"📏 Taille du fichier : {taille} octets")
            else:
                print("⚠️ Attention : Le fichier n'a pas été trouvé après création")
            
            print(f"\n📊 Statistiques :")
            print(f"   - Total lignes : {len(df)}")
            print(f"   - Produits uniques : {df['Produit'].nunique()}")
            print(f"   - Composants uniques : {df['Code Composant'].nunique()}")
            print(f"\n📑 Répartition par étape :")
            for stage, count in df['Gamme'].value_counts().items():
                print(f"   - {stage} : {count} composants")
                
        except PermissionError:
            print(f"❌ Erreur de permission : Le fichier {output_file} est peut-être ouvert dans Excel.")
            print("   Veuillez fermer le fichier s'il est ouvert et réessayer.")
        except Exception as e:
            print(f"❌ Erreur lors de l'écriture du fichier : {str(e)}")
            traceback.print_exc()
            
    except Exception as e:
        print(f"❌ Erreur générale : {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    main()