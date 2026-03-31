import openpyxl
from openpyxl import Workbook

# Définition des étapes (incluant le laquage)
STAGES = ["découpe", "placage de chant", "perçage", "laquage", "montage", "conditionnement"]

def get_stage(description):
    """Retourne l'étape en fonction de la description du composant."""
    desc = str(description).upper() if description else ""
    
    # 1. Découpe : panneaux
    if "PANNEAU" in desc:
        return "découpe"
    
    # 2. Placage de chant : bandes de chant ou colle de chant
    if "BANDE DE CHANT" in desc or "COLLE DE CHANT" in desc:
        return "placage de chant"
    
    # 3. Laquage : produits de peinture, laque, vernis, diluants, catalyseurs, etc.
    laquage_keywords = [
        "DILUANT", "LIQUIDE", "CATALYSEUR", "FINITION", "ISOLANT POLYURETHANNE",
        "FOND BLANC", "PEINTURE", "LAQUE", "VERNIS", "RAL", "TOPCOAT", "THINNER",
        "MEDIUM", "THINNER", "POLYRETHANE", "TOP COAT"
    ]
    if any(kw in desc for kw in laquage_keywords):
        return "laquage"
    
    # 4. Conditionnement : emballages, poignées, vis de poignée, palettes, agrafes
    packaging = [
        "SACHET", "POIGNEE", "CARTON D'EMBALLAGE", "DOUBLE HOOK",
        "VIS POIGNEE", "PALETTE", "AGRAFES"
    ]
    if any(kw in desc for kw in packaging):
        return "conditionnement"
    
    # 5. Par défaut, tout le reste va en montage
    return "montage"

def process_sheet(wb, sheet_name, output_suffix):
    """Traite une feuille et génère un fichier Excel de sortie."""
    if sheet_name not in wb.sheetnames:
        print(f"Feuille '{sheet_name}' introuvable. Ignorée.")
        return None
    
    sheet = wb[sheet_name]
    rows_data = []
    current_product = None
    
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 6:
            continue
        ref, designation, composant, desc_composant, quantity, unit = row[:6]
        
        # Mise à jour du produit courant
        if designation:
            current_product = str(designation).strip()
        
        if not current_product:
            continue
        
        # Ignorer les lignes MOD / ZMOD
        if composant and str(composant).upper() in ["MOD", "ZMOD"]:
            continue
        
        if not desc_composant:
            continue
        
        stage = get_stage(desc_composant)
        
        rows_data.append([
            current_product,
            stage,
            str(desc_composant),
            quantity,
            unit if unit else ""
        ])
    
    # Création du fichier Excel de sortie
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Nomenclature par étape"
    headers = ["Produit", "Gamme", "Désignation Componsant", "Quantité", "Unité Qté"]
    out_ws.append(headers)
    for row in rows_data:
        out_ws.append(row)
    
    output_file = f"Nomenclature_{sheet_name}_par_etape.xlsx"
    out_wb.save(output_file)
    print(f"Fichier {output_file} généré avec {len(rows_data)} lignes (feuille '{sheet_name}').")
    return output_file

def main():
    # Charger le classeur source
    wb = openpyxl.load_workbook("Nomenclature CEJID 2026.xlsx", data_only=True)
    
    # Choisir ici la feuille à traiter : "LAQUEE 2026" ou "LAQUEE BARDAGE"
    target_sheet = "LAQUEE 2026"   # <--- Modifiez selon vos besoins
    
    # Traitement
    process_sheet(wb, target_sheet, "")

if __name__ == "__main__":
    main()