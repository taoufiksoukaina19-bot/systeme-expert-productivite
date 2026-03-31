import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import sys
import io

# Forcer l'encodage UTF-8 pour la console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

class TraçabiliteProduction:
    def __init__(self):
        self.produits = {}
        self.composants = {}
        self.gammes = {}
        self.initialiser_donnees()
    
    def initialiser_donnees(self):
        """Initialise les donnees de production"""
        
        # Definition des produits avec leurs niveaux (P0 = matiere premiere, PF = produit fini)
        self.produits = {
            'MEU050NMABLB': {
                'nom': 'MEUBLE INMAA D 50*36 CM BLANC BRILLANT',
                'type': 'PF',
                'niveau': 3,
                'composants': [
                    {'code': 'MPPAF16SR209', 'quantite': 0.54, 'unite': 'M2'},
                    {'code': 'MPPAF28516BLS', 'quantite': 0.25, 'unite': 'M2'},
                    {'code': 'QUCOCOTHERM0', 'quantite': 0.13, 'unite': 'KG'},
                    {'code': 'QUDVCHPL0810', 'quantite': 4, 'unite': 'U'},
                    {'code': 'QUPOCH500500', 'quantite': 2, 'unite': 'U'},
                    {'code': 'QUFCHAFR3500', 'quantite': 2, 'unite': 'U'},
                    {'code': 'QUFCHNOR3500', 'quantite': 2, 'unite': 'U'},
                    {'code': 'MPBC819BL209', 'quantite': 5.66, 'unite': 'ML'},
                    {'code': 'QUVITOU08300', 'quantite': 12, 'unite': 'U'},
                    {'code': 'QUVICH041600', 'quantite': 8, 'unite': 'U'}
                ]
            },
            'MEU6036K2BLG': {
                'nom': 'MEUBLE K2 D 60*36 CM BLANC GLOSS',
                'type': 'PF',
                'niveau': 3,
                'composants': [
                    {'code': 'MPPAF16SR209', 'quantite': 0.63, 'unite': 'M2'},
                    {'code': 'MPPAF28516BLS', 'quantite': 0.29, 'unite': 'M2'},
                    {'code': 'QUCOCOTHERM0', 'quantite': 0.13, 'unite': 'KG'},
                    {'code': 'QUDVCHPL0810', 'quantite': 4, 'unite': 'U'},
                    {'code': 'QUPO12906422', 'quantite': 2, 'unite': 'U'},
                    {'code': 'QUFCHAFR3500', 'quantite': 2, 'unite': 'U'},
                    {'code': 'QUFCHNOR3500', 'quantite': 2, 'unite': 'U'},
                    {'code': 'MPBC819BL209', 'quantite': 6.29, 'unite': 'ML'},
                    {'code': 'QUVITOU08300', 'quantite': 12, 'unite': 'U'},
                    {'code': 'QUVICH041600', 'quantite': 8, 'unite': 'U'}
                ]
            },
            'MEU060DENOAK': {
                'nom': 'MEUBLE DENIA D 60 CM OAK',
                'type': 'PF',
                'niveau': 3,
                'composants': [
                    {'code': 'MPPAF28516COS', 'quantite': 1.19, 'unite': 'M2'},
                    {'code': 'QUCOCOTHERM0', 'quantite': 0.13, 'unite': 'KG'},
                    {'code': 'ACTIRG250816', 'quantite': 0.36, 'unite': 'ML'},
                    {'code': 'ACTIRG198150', 'quantite': 1.9, 'unite': 'ML'},
                    {'code': 'QUDVCHPL0810', 'quantite': 4, 'unite': 'U'},
                    {'code': 'QUSACOFR3500', 'quantite': 1, 'unite': 'U'},
                    {'code': 'QUPOCH0963AB', 'quantite': 1, 'unite': 'U'},
                    {'code': 'MPBC0819N822', 'quantite': 13, 'unite': 'ML'},
                    {'code': 'QUVITOU08300', 'quantite': 20, 'unite': 'U'},
                    {'code': 'QUVICH041600', 'quantite': 8, 'unite': 'U'}
                ]
            },
            'MEU060MADOAK': {
                'nom': 'MEUBLE MADRID D 60 CM OAK',
                'type': 'PF',
                'niveau': 3,
                'composants': [
                    {'code': 'MPPAF28516COS', 'quantite': 1.38, 'unite': 'M2'},
                    {'code': 'QUCOCOTHERM0', 'quantite': 0.13, 'unite': 'KG'},
                    {'code': 'ACTIRG250816', 'quantite': 0.36, 'unite': 'ML'},
                    {'code': 'ACTIRG198150', 'quantite': 1.7, 'unite': 'ML'},
                    {'code': 'ACTIRG198860', 'quantite': 1.22, 'unite': 'ML'},
                    {'code': 'QUDVCHPL0810', 'quantite': 4, 'unite': 'U'},
                    {'code': 'QUSACOFR3500', 'quantite': 2, 'unite': 'U'},
                    {'code': 'QUPOCH0963AB', 'quantite': 2, 'unite': 'U'},
                    {'code': 'MPBC0819N822', 'quantite': 11.55, 'unite': 'ML'},
                    {'code': 'QUVITOU08300', 'quantite': 20, 'unite': 'U'},
                    {'code': 'QUVICH041600', 'quantite': 8, 'unite': 'U'}
                ]
            },
            'MEU080MADOAK': {
                'nom': 'MEUBLE MADRID D 80 CM OAK',
                'type': 'PF',
                'niveau': 3,
                'composants': [
                    {'code': 'MPPAF28516COS', 'quantite': 1.62, 'unite': 'M2'},
                    {'code': 'QUCOCOTHERM0', 'quantite': 0.13, 'unite': 'KG'},
                    {'code': 'ACTIRG250816', 'quantite': 0.55, 'unite': 'ML'},
                    {'code': 'ACTIRG198150', 'quantite': 2.07, 'unite': 'ML'},
                    {'code': 'ACTIRG198860', 'quantite': 1.38, 'unite': 'ML'},
                    {'code': 'QUDVCHPL0810', 'quantite': 4, 'unite': 'U'},
                    {'code': 'QUSACOFR3500', 'quantite': 2, 'unite': 'U'},
                    {'code': 'QUPOCH704000', 'quantite': 2, 'unite': 'U'},
                    {'code': 'MPBC0819N822', 'quantite': 13.39, 'unite': 'ML'},
                    {'code': 'QUVITOU08300', 'quantite': 20, 'unite': 'U'},
                    {'code': 'QUVICH041600', 'quantite': 8, 'unite': 'U'}
                ]
            },
            'MEU060SEBPBLB': {
                'nom': 'MEUBLE SEBOU (P) 60 CM BLANC BRILLANT',
                'type': 'PF',
                'niveau': 3,
                'composants': [
                    {'code': 'MPPAF16SR209', 'quantite': 1.14, 'unite': 'M2'},
                    {'code': 'MPPAF28516BLS', 'quantite': 0.67, 'unite': 'M2'},
                    {'code': 'QUCOCOTHERM0', 'quantite': 0.13, 'unite': 'KG'},
                    {'code': 'ACSALB4012AL', 'quantite': 4, 'unite': 'U'},
                    {'code': 'ACPOSMPPL10A', 'quantite': 2, 'unite': 'U'},
                    {'code': 'QUFCHAFR3500', 'quantite': 2, 'unite': 'U'},
                    {'code': 'QUFCHNOR3500', 'quantite': 2, 'unite': 'U'},
                    {'code': 'MPBC819BL209', 'quantite': 8.67, 'unite': 'ML'},
                    {'code': 'MPBC8192011N', 'quantite': 5.73, 'unite': 'ML'},
                    {'code': 'QUVITOU08300', 'quantite': 20, 'unite': 'U'},
                    {'code': 'QUVICH041600', 'quantite': 24, 'unite': 'U'}
                ]
            }
        }
        
        # Definition des composants (P0 = matiere premiere)
        self.composants = {
            'MPPAF16SR209': {
                'nom': 'PANNEAU FINSA 275*210*16 SUPER DECOR 2F WHITE SR209',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Panneau',
                'fournisseur': 'Finsa',
                'stock': 500,
                'unite': 'M2'
            },
            'MPPAF28516BLS': {
                'nom': 'PANNEAU FINSA 285*210*16 SUPER DECOR 2F BLANCO STANDARD',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Panneau',
                'fournisseur': 'Finsa',
                'stock': 500,
                'unite': 'M2'
            },
            'MPPAF28516COS': {
                'nom': 'PANNEAU FINSA 285*210*16 SUPER DECOR 2F CAMBRIAN OAK SEGA',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Panneau',
                'fournisseur': 'Finsa',
                'stock': 400,
                'unite': 'M2'
            },
            'QUCOCOTHERM0': {
                'nom': 'COLLE DE CHANT RAYT MA 6520 EN 25KG',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Colle',
                'fournisseur': 'Rayt',
                'stock': 100,
                'unite': 'KG'
            },
            'QUDVCHPL0810': {
                'nom': 'CHEVILLE PLASTIQUE D 08*10',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Quincaillerie',
                'fournisseur': 'Standard',
                'stock': 10000,
                'unite': 'U'
            },
            'QUFCHAFR3500': {
                'nom': 'CHARNIERE A FREIN 35',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Charniere',
                'fournisseur': 'Hettich',
                'stock': 2000,
                'unite': 'U'
            },
            'QUFCHNOR3500': {
                'nom': 'CHARNIERE NORMAL D 35',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Charniere',
                'fournisseur': 'Hettich',
                'stock': 2000,
                'unite': 'U'
            },
            'QUVITOU08300': {
                'nom': 'TOURILLON 08*30',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Quincaillerie',
                'fournisseur': 'Standard',
                'stock': 5000,
                'unite': 'U'
            },
            'QUVICH041600': {
                'nom': 'VIS SPAX 4*16 CHROME',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Vis',
                'fournisseur': 'Spax',
                'stock': 10000,
                'unite': 'U'
            },
            'MPBC819BL209': {
                'nom': 'BANDE DE CHANT PROTEC 08*19 BLANCO BRILLANT',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Chant',
                'fournisseur': 'Protec',
                'stock': 1000,
                'unite': 'ML'
            },
            'MPBC8192011N': {
                'nom': 'BANDE DE CHANT PROTEC 0.8*19 2011N BLANCO SERF',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Chant',
                'fournisseur': 'Protec',
                'stock': 1000,
                'unite': 'ML'
            },
            'MPBC0819N822': {
                'nom': 'BANDE DE CHANT PROTEC 08*19 ROBLE PORO OAK SEGA',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Chant',
                'fournisseur': 'Protec',
                'stock': 800,
                'unite': 'ML'
            },
            'QUPOCH500500': {
                'nom': 'POIGNEE JRG CHROME 50-050-03',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Poignee',
                'fournisseur': 'JRG',
                'stock': 1500,
                'unite': 'U'
            },
            'QUPO12906422': {
                'nom': 'POIGNEE JRG 70-129-064-03B',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Poignee',
                'fournisseur': 'JRG',
                'stock': 1500,
                'unite': 'U'
            },
            'QUPOCH0963AB': {
                'nom': 'POIGNEE JRG CHROME 70-096-300-96-89E',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Poignee',
                'fournisseur': 'JRG',
                'stock': 1500,
                'unite': 'U'
            },
            'QUPOCH704000': {
                'nom': 'POIGNEE JRG CHROME 70-096-400-96-89E',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Poignee',
                'fournisseur': 'JRG',
                'stock': 1500,
                'unite': 'U'
            },
            'QUSACOFR3500': {
                'nom': 'COULISSE A FREIN D 35',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Coulisse',
                'fournisseur': 'Standard',
                'stock': 1000,
                'unite': 'U'
            },
            'ACTIRG250816': {
                'nom': 'TIRAGUARDERA DE 2050*86*16 GRIS GU',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Accessoire',
                'fournisseur': 'Standard',
                'stock': 500,
                'unite': 'ML'
            },
            'ACTIRG198150': {
                'nom': 'TIRAGUARDERA DE 1980*150*12 GRIS GU',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Accessoire',
                'fournisseur': 'Standard',
                'stock': 500,
                'unite': 'ML'
            },
            'ACTIRG198860': {
                'nom': 'TIRAGUARDERA DE 1980*86*12 GRIS GU',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Accessoire',
                'fournisseur': 'Standard',
                'stock': 500,
                'unite': 'ML'
            },
            'ACSALB4012AL': {
                'nom': 'PATA ALBUFERA 120 mm ALUMINIO',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Accessoire',
                'fournisseur': 'Standard',
                'stock': 1000,
                'unite': 'U'
            },
            'ACPOSMPPL10A': {
                'nom': 'POMO SAMB PPL10 100*100 ALUMINIO',
                'type': 'P0',
                'niveau': 0,
                'categorie': 'Accessoire',
                'fournisseur': 'Standard',
                'stock': 500,
                'unite': 'U'
            }
        }
        
        # Definition des gammes operatoires par produit
        self.gammes = {
            'MEU050NMABLB': [
                {'poste': 'Decoupe', 'operation': 'Decoupe panneaux MDF', 'machine': 'Scie a panneaux', 'temps': 15, 'sequence': 1},
                {'poste': 'Usinage', 'operation': 'Perçage des trous', 'machine': 'Centre usinage', 'temps': 12, 'sequence': 2},
                {'poste': 'Chantonnage', 'operation': 'Application chants', 'machine': 'Chanteuse', 'temps': 10, 'sequence': 3},
                {'poste': 'Assemblage', 'operation': 'Montage quincaillerie', 'machine': 'Poste assemblage', 'temps': 20, 'sequence': 4},
                {'poste': 'Controle', 'operation': 'Controle qualite', 'machine': 'Poste controle', 'temps': 8, 'sequence': 5},
                {'poste': 'Emballage', 'operation': 'Emballage final', 'machine': 'Poste emballage', 'temps': 10, 'sequence': 6}
            ],
            'MEU6036K2BLG': [
                {'poste': 'Decoupe', 'operation': 'Decoupe panneaux MDF', 'machine': 'Scie a panneaux', 'temps': 18, 'sequence': 1},
                {'poste': 'Usinage', 'operation': 'Perçage des trous', 'machine': 'Centre usinage', 'temps': 14, 'sequence': 2},
                {'poste': 'Chantonnage', 'operation': 'Application chants', 'machine': 'Chanteuse', 'temps': 12, 'sequence': 3},
                {'poste': 'Assemblage', 'operation': 'Montage quincaillerie', 'machine': 'Poste assemblage', 'temps': 25, 'sequence': 4},
                {'poste': 'Controle', 'operation': 'Controle qualite', 'machine': 'Poste controle', 'temps': 10, 'sequence': 5},
                {'poste': 'Emballage', 'operation': 'Emballage final', 'machine': 'Poste emballage', 'temps': 12, 'sequence': 6}
            ],
            'MEU060DENOAK': [
                {'poste': 'Decoupe', 'operation': 'Decoupe panneaux melamine', 'machine': 'Scie a panneaux', 'temps': 20, 'sequence': 1},
                {'poste': 'Usinage', 'operation': 'Perçage + fraisage', 'machine': 'Centre usinage', 'temps': 18, 'sequence': 2},
                {'poste': 'Chantonnage', 'operation': 'Application chants chene', 'machine': 'Chanteuse', 'temps': 15, 'sequence': 3},
                {'poste': 'Assemblage', 'operation': 'Montage tiroirs', 'machine': 'Poste assemblage', 'temps': 30, 'sequence': 4},
                {'poste': 'Controle', 'operation': 'Controle qualite', 'machine': 'Poste controle', 'temps': 12, 'sequence': 5},
                {'poste': 'Emballage', 'operation': 'Emballage final', 'machine': 'Poste emballage', 'temps': 12, 'sequence': 6}
            ],
            'MEU060MADOAK': [
                {'poste': 'Decoupe', 'operation': 'Decoupe panneaux melamine', 'machine': 'Scie a panneaux', 'temps': 22, 'sequence': 1},
                {'poste': 'Usinage', 'operation': 'Perçage + fraisage', 'machine': 'Centre usinage', 'temps': 20, 'sequence': 2},
                {'poste': 'Chantonnage', 'operation': 'Application chants chene', 'machine': 'Chanteuse', 'temps': 16, 'sequence': 3},
                {'poste': 'Assemblage', 'operation': 'Montage 2 tiroirs', 'machine': 'Poste assemblage', 'temps': 35, 'sequence': 4},
                {'poste': 'Controle', 'operation': 'Controle qualite', 'machine': 'Poste controle', 'temps': 12, 'sequence': 5},
                {'poste': 'Emballage', 'operation': 'Emballage final', 'machine': 'Poste emballage', 'temps': 12, 'sequence': 6}
            ],
            'MEU080MADOAK': [
                {'poste': 'Decoupe', 'operation': 'Decoupe panneaux melamine', 'machine': 'Scie a panneaux', 'temps': 25, 'sequence': 1},
                {'poste': 'Usinage', 'operation': 'Perçage + fraisage', 'machine': 'Centre usinage', 'temps': 22, 'sequence': 2},
                {'poste': 'Chantonnage', 'operation': 'Application chants chene', 'machine': 'Chanteuse', 'temps': 18, 'sequence': 3},
                {'poste': 'Assemblage', 'operation': 'Montage 2 tiroirs', 'machine': 'Poste assemblage', 'temps': 38, 'sequence': 4},
                {'poste': 'Controle', 'operation': 'Controle qualite', 'machine': 'Poste controle', 'temps': 15, 'sequence': 5},
                {'poste': 'Emballage', 'operation': 'Emballage final', 'machine': 'Poste emballage', 'temps': 15, 'sequence': 6}
            ],
            'MEU060SEBPBLB': [
                {'poste': 'Decoupe', 'operation': 'Decoupe panneaux MDF', 'machine': 'Scie a panneaux', 'temps': 20, 'sequence': 1},
                {'poste': 'Usinage', 'operation': 'Perçage des trous', 'machine': 'Centre usinage', 'temps': 15, 'sequence': 2},
                {'poste': 'Chantonnage', 'operation': 'Application chants', 'machine': 'Chanteuse', 'temps': 14, 'sequence': 3},
                {'poste': 'Assemblage', 'operation': 'Montage quincaillerie', 'machine': 'Poste assemblage', 'temps': 25, 'sequence': 4},
                {'poste': 'Controle', 'operation': 'Controle qualite', 'machine': 'Poste controle', 'temps': 10, 'sequence': 5},
                {'poste': 'Emballage', 'operation': 'Emballage final', 'machine': 'Poste emballage', 'temps': 12, 'sequence': 6}
            ]
        }
    
    def tracer_produit(self, code_produit):
        """Trace completement un produit depuis P0 jusqu'a PF"""
        if code_produit not in self.produits:
            return f"Produit {code_produit} non trouve"
        
        trace = {
            'produit': code_produit,
            'nom': self.produits[code_produit]['nom'],
            'niveau_final': self.produits[code_produit]['niveau'],
            'composants_trace': [],
            'gamme_operatoire': self.gammes.get(code_produit, []),
            'recapitulatif_composants': defaultdict(lambda: {'quantite': 0, 'unite': '', 'categorie': ''})
        }
        
        for composant in self.produits[code_produit]['composants']:
            trace_composant = self._tracer_composant(composant['code'], composant['quantite'], composant.get('unite', 'U'))
            trace['composants_trace'].append(trace_composant)
            
            code = composant['code']
            trace['recapitulatif_composants'][code]['quantite'] += composant['quantite']
            trace['recapitulatif_composants'][code]['unite'] = composant.get('unite', 'U')
            if code in self.composants:
                trace['recapitulatif_composants'][code]['categorie'] = self.composants[code]['categorie']
        
        return trace
    
    def _tracer_composant(self, code_composant, quantite, unite='U'):
        """Trace un composant individuel jusqu'a sa matiere premiere"""
        if code_composant in self.composants:
            composant_info = self.composants[code_composant]
            return {
                'code': code_composant,
                'nom': composant_info['nom'],
                'type': composant_info['type'],
                'niveau': composant_info['niveau'],
                'quantite': quantite,
                'unite': composant_info.get('unite', unite),
                'categorie': composant_info.get('categorie', 'N/A'),
                'fournisseur': composant_info.get('fournisseur', 'N/A')
            }
        else:
            return {
                'code': code_composant,
                'nom': 'Composant non reference',
                'type': 'Inconnu',
                'niveau': -1,
                'quantite': quantite,
                'unite': unite,
                'categorie': 'Inconnu',
                'fournisseur': 'Inconnu'
            }
    
    def calculer_besoin_matieres(self, code_produit, quantite_produite=1):
        """Calcule les besoins en matieres premieres pour une quantite produite"""
        trace = self.tracer_produit(code_produit)
        
        besoins = defaultdict(lambda: {'quantite': 0, 'unite': '', 'categorie': '', 'fournisseur': ''})
        
        for composant in trace['composants_trace']:
            if composant['type'] == 'P0':
                code = composant['code']
                besoins[code]['quantite'] += composant['quantite'] * quantite_produite
                besoins[code]['unite'] = composant['unite']
                besoins[code]['categorie'] = composant['categorie']
                besoins[code]['fournisseur'] = composant['fournisseur']
        
        return besoins
    
    def generer_fiche_production(self, code_produit, quantite=1):
        """Genere une fiche de production complete"""
        trace = self.tracer_produit(code_produit)
        
        fiche = {
            'reference': code_produit,
            'designation': trace['nom'],
            'quantite_a_produire': quantite,
            'nomenclature': [],
            'besoins_matieres': [],
            'gamme': trace['gamme_operatoire'],
            'temps_total': sum(op['temps'] for op in trace['gamme_operatoire']) * quantite
        }
        
        for composant in trace['composants_trace']:
            fiche['nomenclature'].append({
                'Code': composant['code'],
                'Designation': composant['nom'],
                'Type': composant['type'],
                'Niveau': composant['niveau'],
                'Quantite': composant['quantite'] * quantite,
                'Unite': composant['unite']
            })
        
        besoins = self.calculer_besoin_matieres(code_produit, quantite)
        for code, info in besoins.items():
            fiche['besoins_matieres'].append({
                'Code': code,
                'Designation': self.composants.get(code, {}).get('nom', ''),
                'Quantite': info['quantite'],
                'Unite': info['unite'],
                'Categorie': info['categorie'],
                'Fournisseur': info['fournisseur']
            })
        
        return fiche
    
    def generer_excel(self, output_file='Traçabilite_Production.xlsx'):
        """Genere le fichier Excel complet"""
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            produits_pf = []
            for code, info in self.produits.items():
                if info['type'] == 'PF':
                    produits_pf.append({
                        'Reference': code,
                        'Designation': info['nom'],
                        'Niveau': info['niveau'],
                        'Nb Composants': len(info['composants'])
                    })
            df_produits = pd.DataFrame(produits_pf)
            df_produits.to_excel(writer, sheet_name='Produits_Finis', index=False)
            
            matieres_p0 = []
            for code, info in self.composants.items():
                if info['type'] == 'P0':
                    matieres_p0.append({
                        'Code': code,
                        'Designation': info['nom'],
                        'Categorie': info['categorie'],
                        'Fournisseur': info['fournisseur'],
                        'Stock': info['stock'],
                        'Unite': info.get('unite', '')
                    })
            df_p0 = pd.DataFrame(matieres_p0)
            df_p0.to_excel(writer, sheet_name='Matieres_Premieres_P0', index=False)
            
            for code in self.produits.keys():
                fiche = self.generer_fiche_production(code)
                
                df_nom = pd.DataFrame(fiche['nomenclature'])
                if not df_nom.empty:
                    sheet_name = f"Trace_{code}"[:31]
                    df_nom.to_excel(writer, sheet_name=sheet_name, index=False)
                
                df_besoins = pd.DataFrame(fiche['besoins_matieres'])
                if not df_besoins.empty:
                    sheet_name = f"Besoins_{code}"[:31]
                    df_besoins.to_excel(writer, sheet_name=sheet_name, index=False)
            
            gammes_data = []
            for code, gamme in self.gammes.items():
                for op in gamme:
                    gammes_data.append({
                        'Produit': code,
                        'Sequence': op['sequence'],
                        'Poste': op['poste'],
                        'Operation': op['operation'],
                        'Machine': op['machine'],
                        'Temps (min)': op['temps']
                    })
            df_gammes = pd.DataFrame(gammes_data)
            df_gammes.to_excel(writer, sheet_name='Gammes_Operatoires', index=False)
            
            synthese = []
            for code in self.produits.keys():
                besoins = self.calculer_besoin_matieres(code)
                for comp_code, info in besoins.items():
                    synthese.append({
                        'Produit': code,
                        'Composant': comp_code,
                        'Quantite': info['quantite'],
                        'Unite': info['unite'],
                        'Categorie': info['categorie']
                    })
            df_synthese = pd.DataFrame(synthese)
            df_synthese.to_excel(writer, sheet_name='Synthese_Besoins', index=False)
            
            temps_production = []
            for code in self.produits.keys():
                fiche = self.generer_fiche_production(code)
                temps_production.append({
                    'Produit': code,
                    'Designation': fiche['designation'],
                    'Temps_Total_Unitaire': sum(op['temps'] for op in fiche['gamme']),
                    'Nb_Operations': len(fiche['gamme']),
                    'Nb_Composants': len(fiche['nomenclature'])
                })
            df_temps = pd.DataFrame(temps_production)
            df_temps.to_excel(writer, sheet_name='Temps_Production', index=False)
        
        self._format_excel(output_file)
        print(f"[OK] Fichier genere: {output_file}")
    
    def _format_excel(self, filename):
        """Applique un formatage professionnel"""
        wb = load_workbook(filename)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for column in ws.columns:
                    max_length = 0
                    col_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
        
        wb.save(filename)
    
    def rechercher_produit(self, code_produit):
        """Recherche et affiche la tracabilite complete d'un produit"""
        trace = self.tracer_produit(code_produit)
        
        print(f"\n{'='*80}")
        print(f"TRACABILITE COMPLETE DU PRODUIT: {code_produit}")
        print(f"{'='*80}")
        print(f"Nom: {trace['nom']}")
        print(f"Niveau final: P{trace['niveau_final']} (Produit Fini)")
        print(f"\n--- COMPOSANTS (Du PF vers P0) ---")
        
        for comp in trace['composants_trace']:
            niveau_str = "P0 (Matiere premiere)" if comp['niveau'] == 0 else f"P{comp['niveau']}"
            print(f"  {comp['code']}: {comp['nom'][:50]}...")
            print(f"    Quantite: {comp['quantite']} {comp['unite']} | Type: {niveau_str} | Categorie: {comp['categorie']}")
        
        print(f"\n--- GAMME OPERATOIRE ---")
        for op in trace['gamme_operatoire']:
            print(f"  {op['sequence']}. {op['poste']} - {op['operation']}")
            print(f"     Machine: {op['machine']} | Temps: {op['temps']} min")
        
        return trace


def main():
    systeme = TraçabiliteProduction()
    
    print("PRODUITS DISPONIBLES:")
    print("-" * 60)
    for code, info in systeme.produits.items():
        if info['type'] == 'PF':
            print(f"  {code}: {info['nom']}")
    
    produit_test = 'MEU050NMABLB'
    systeme.rechercher_produit(produit_test)
    
    print(f"\n{'='*80}")
    print(f"BESOINS EN MATIERES PREMIERES POUR 10 UNITES DE {produit_test}")
    print(f"{'='*80}")
    
    besoins = systeme.calculer_besoin_matieres(produit_test, quantite_produite=10)
    for code, info in besoins.items():
        print(f"  {code}: {info['quantite']} {info['unite']} - {info['categorie']}")
    
    print(f"\n{'='*80}")
    print("GENERATION DU FICHIER EXCEL...")
    print(f"{'='*80}")
    systeme.generer_excel('Traçabilite_Production_CEJID.xlsx')
    
    print("\n[SUCCES] Fichier genere avec succes!")
    print("Le fichier contient:")
    print("  - Liste des produits finis (PF)")
    print("  - Matieres premieres (P0)")
    print("  - Tracabilite complete par produit")
    print("  - Gammes operatoires")
    print("  - Synthese des besoins")
    print("  - Temps de production")


if __name__ == "__main__":
    main()