import openpyxl
import pandas as pd

# Définition des étapes
STAGES = ["découpe", "placage de chant", "perçage", "montage", "conditionnement"]

def get_stage(description):
    """Retourne l'étape à partir de la description du composant."""
    desc = str(description).upper() if description else ""
    if "PANNEAU" in desc:
        return "découpe"
    if "BANDE DE CHANT" in desc or "COLLE DE CHANT" in desc:
        return "placage de chant"
    # Mots clés pour le conditionnement (emballage)
    packaging = ["SACHET", "POIGNEE", "CARTON D'EMBALLAGE", "DOUBLE HOOK",
                 "VIS POIGNEE", "PALETTE", "AGRAFES"]
    if any(kw in desc for kw in packaging):
        return "conditionnement"
    # Par défaut, tout le reste va en montage
    return "montage"

def main():
    # Charger le classeur source avec évaluation des formules
    wb = openpyxl.load_workbook("Nomenclature CEJID 2026.xlsx", data_only=True)

    all_rows = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        current_product = None

        for row in sheet.iter_rows(values_only=True):
            if len(row) < 6:
                continue
            # Colonnes : Reference, Designation, Composant, Désignation Componsant, Quantité, Unité
            ref, designation, composant, desc_composant, quantity, unit = row[:6]

            # Mise à jour du produit courant (propagation)
            if designation:
                current_product = str(designation).strip()

            # Ignorer les lignes sans produit
            if not current_product:
                continue

            # Ignorer les composants MOD / ZMOD
            if composant and str(composant).upper() in ["MOD", "ZMOD"]:
                continue

            # Ignorer les lignes sans description de composant
            if not desc_composant:
                continue

            # Déterminer l'étape (gamme)
            stage = get_stage(desc_composant)

            all_rows.append({
                "Produit": current_product,
                "Gamme": stage,
                "Désignation Componsant": str(desc_composant),
                "Quantité": quantity,
                "Unité Qté": unit if unit else ""
            })

    # Créer le DataFrame et l'exporter
    df = pd.DataFrame(all_rows)
    output_file = "Nomenclature_par_etape.xlsx"
    try:
        df.to_excel(output_file, index=False)
        print(f"Fichier {output_file} généré avec {len(df)} lignes.")
    except PermissionError:
        print(f"Erreur : le fichier {output_file} est ouvert. Veuillez le fermer et relancer.")

if __name__ == "__main__":
    main()