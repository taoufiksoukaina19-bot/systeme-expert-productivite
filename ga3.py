import openpyxl
from openpyxl import Workbook

# Définition des étapes (incluant le laquage)
STAGES = ["découpe", "placage de chant", "perçage", "laquage", "montage", "conditionnement"]

def get_stage(description):
    """Retourne l'étape en fonction de la description du composant."""
    desc = str(description).upper() if description else ""
    
    # 1. Découpe : panneaux
    if "PANNEAU" in desc:
        return "découpe"
    
    # 2. Placage de chant : bandes de chant ou colle de chant
    if "BANDE DE CHANT" in desc or "COLLE DE CHANT" in desc:
        return "placage de chant"
    
    # 3. Laquage : produits de peinture, laque, vernis, diluants, catalyseurs, etc.
    laquage_keywords = [
        "DILUANT", "LIQUIDE", "CATALYSEUR", "FINITION", "ISOLANT POLYURETHANNE",
        "FOND BLANC", "PEINTURE", "LAQUE", "VERNIS", "RAL", "TOPCOAT", "THINNER",
        "MEDIUM", "POLYRETHANE"
    ]
    if any(kw in desc for kw in laquage_keywords):
        return "laquage"
    
    # 4. Conditionnement : emballages, poignées, vis de poignée, palettes, agrafes
    packaging = [
        "SACHET", "POIGNEE", "CARTON D'EMBALLAGE", "DOUBLE HOOK",
        "VIS POIGNEE", "PALETTE", "AGRAFES"
    ]
    if any(kw in desc for kw in packaging):
        return "conditionnement"
    
    # 5. Par défaut, tout le reste va en montage
    return "montage"

def main():
    # Charger le classeur source (évaluer les formules)
    wb = openpyxl.load_workbook("Nomenclature CEJID 2026.xlsx", data_only=True)
    
    # Cibler la feuille 4
    target_sheet = "LAQUEE BARDAGE"
    if target_sheet not in wb.sheetnames:
        print(f"Feuille '{target_sheet}' introuvable. Feuilles disponibles : {wb.sheetnames}")
        return
    
    sheet = wb[target_sheet]
    rows_data = []
    current_product = None
    
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 6:
            continue
        ref, designation, composant, desc_composant, quantity, unit = row[:6]
        
        # Mise à jour du produit courant (propagation vers le bas)
        if designation:
            current_product = str(designation).strip()
        
        if not current_product:
            continue
        
        # Ignorer les lignes MOD / ZMOD
        if composant and str(composant).upper() in ["MOD", "ZMOD"]:
            continue
        
        if not desc_composant:
            continue
        
        # Déterminer l'étape
        stage = get_stage(desc_composant)
        
        rows_data.append([
            current_product,
            stage,
            str(desc_composant),
            quantity,
            unit if unit else ""
        ])
    
    # Création du fichier Excel de sortie
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Nomenclature par étape"
    headers = ["Produit", "Gamme", "Désignation Componsant", "Quantité", "Unité Qté"]
    out_ws.append(headers)
    for row in rows_data:
        out_ws.append(row)
    
    output_file = "Nomenclature_LAQUEE_BARDAGE_par_etape.xlsx"
    out_wb.save(output_file)
    print(f"Fichier {output_file} généré avec {len(rows_data)} lignes (feuille '{target_sheet}').")

if __name__ == "__main__":
    main()